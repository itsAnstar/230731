import os
import json
import requests
import re
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import openpyxl
import threading

# 弹出文件选择窗口
excel_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])

if not excel_path:
    print("未选择Excel文件")
    exit()

# 打开Excel文件
workbook = openpyxl.load_workbook(excel_path)
sheet = workbook.active

# 定义变量
flink_column = 1  # 第一列
date_column = 2   # 第二列
author_column = 3 # 第三列

# 读取 Excel 数据
download_tasks = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    flink = row[flink_column - 1]
    date = row[date_column - 1]
    author = row[author_column - 1]
    video_info = f"{date}_{author}"

    download_tasks.append((flink, video_info))

# 获取用户选择的保存目录
root = tk.Tk()
root.withdraw()  # 隐藏主窗口
save_dir = filedialog.askdirectory(title="选择保存目录")
root.destroy()  # 关闭隐藏的主窗口

if not save_dir:
    print("用户取消选择保存目录！")
    exit()

# 定义一个函数来处理下载任务
def download_video(task):
    flink, video_info = task

    url = flink
    headers = {
        'cookie': 'douyin.com; device_web_cpu_core=8; device_web_memory_size=8; webcast_local_quality=null; ttwid=1%7CMGhi6SefHMdUj-2dvOPmQ5zFKo7syPiQf_LdVo_sm9g%7C1691772890%7C900dcd00007c7a54c4bbe58a2c39823d8f1b21e41c12ab840b80faa23f39379b; passport_csrf_token=675830ff4a4ca21e4473ca5989477067; passport_csrf_token_default=675830ff4a4ca21e4473ca5989477067; s_v_web_id=verify_ll6twkdq_3GNpygbG_8yuj_4d9e_82E9_wsUAWUlhDd44; FORCE_LOGIN=%7B%22videoConsumedRemainSeconds%22%3A180%7D; xgplayer_user_id=735486684335; __live_version__=%221.1.1.2573%22; volume_info=%7B%22isUserMute%22%3Afalse%2C%22isMute%22%3Afalse%2C%22volume%22%3A0.5%7D; download_guide=%223%2F20230819%2F1%22; pwa2=%220%7C0%7C3%7C0%22; __ac_nonce=064e7ad1d00ce5d0217d0; __ac_signature=_02B4Z6wo00f01wQ.rzwAAIDAVEjrxIJXIW8EH6uAAKXyLOls23SZWmI1Sr.Thuy2Xy0963jwB8P6yesohoIwm406e9HdbsLwTw3B4dQrA9N98HyM78e348uYsX6fqcKQrtG4W40DmcNSal4072; strategyABtestKey=%221692904734.735%22; stream_recommend_feed_params=%22%7B%5C%22cookie_enabled%5C%22%3Atrue%2C%5C%22screen_width%5C%22%3A2560%2C%5C%22screen_height%5C%22%3A1440%2C%5C%22browser_online%5C%22%3Atrue%2C%5C%22cpu_core_num%5C%22%3A8%2C%5C%22device_memory%5C%22%3A8%2C%5C%22downlink%5C%22%3A10%2C%5C%22effective_type%5C%22%3A%5C%224g%5C%22%2C%5C%22round_trip_time%5C%22%3A0%7D%22; bd_ticket_guard_client_data=eyJiZC10aWNrZXQtZ3VhcmQtdmVyc2lvbiI6MiwiYmQtdGlja2V0LWd1YXJkLWl0ZXJhdGlvbi12ZXJzaW9uIjoxLCJiZC10aWNrZXQtZ3VhcmQtcmVlLXB1YmxpYy1rZXkiOiJCRVJSdFVNdzVjbkdaUXVWa0JaL2NiV2VwZjc4RDRsVjV1N0tCbnBiVTVtMVhUdUQ2WXVneEdrL1piaERTZkpqUCtkd1ZrV3pYZ0dtbHdwZWpxUkNlRkk9IiwiYmQtdGlja2V0LWd1YXJkLXdlYi12ZXJzaW9uIjoxfQ==; home_can_add_dy_2_desktop=%221%22; msToken=dsSVJiV2PkTYD5ZAytr1eYw8JU8YmiZAFxmkp3utfyWEhp5pma8bgTWD3eAQqZXH2o1HTge41xSfaCOQS6mf44v-Vlwm16eneBxhdyJ4eGUAJBwHBHM69EPaFMaNmA==; msToken=VzhzlmgUKvmbt3VTvbXExUHRjH5-01YKlFGua7arcxMI4lRqBHOHSDRoZhmN99Z4lRG1MB91Dqve61CMNZCMw6g3OyjgcWXzCyQamWXV0QJLBwfnIq4Z; tt_scid=WMeBwmdU0zVDaL5127Odic6ksAE2meTszokN1BoOn0zhOYMLzT3B4NSd2mg9glVXae62; IsDouyinActive=false',
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'
    }
    response = requests.get(url=url, headers=headers)
    video_info = re.findall('<script id="RENDER_DATA" type="application/json">(.*?)</script', response.text)[0]
    video_info = requests.utils.unquote(video_info)
    json_data = json.loads(video_info)
    video_url = 'https:' + \
                json_data['74931a6b75e09238f154ab1577c994c9']['aweme']['detail']['video']['bitRateList'][0]['playAddr'][0]['src']

    if response.status_code == 200:
        video_path = os.path.join(save_dir, f"{video_info}.mp4")
        with open(video_path, "wb") as f:
            f.write(response.content)
        print(f"视频下载成功！保存路径：{video_path}")
    else:
        print("视频下载失败！")

    video_url = 'https:' + \
                json_data['74931a6b75e09238f154ab1577c994c9']['aweme']['detail']['video']['bitRateList'][0]['playAddr'][0]['src']

    response = requests.get(video_url)

    if response.status_code == 200:
        video_path = os.path.join(save_dir, f"{video_info}.mp4")
        with open(video_path, "wb") as f:
            f.write(response.content)
        print(f"视频下载成功！保存路径：{video_path}")
    else:
        print("视频下载失败！")

# 创建线程列表
threads = []

# 处理下载任务
for task in download_tasks:
    # 创建线程并启动
    thread = threading.Thread(target=download_video, args=(task,))
    threads.append(thread)
    thread.start()

# 等待所有线程完成
for thread in threads:
    thread.join()

# 关闭Excel文件
workbook.close()
