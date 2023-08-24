import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

def convert_datetime_to_month_day(datetime_str):
    datetime_obj = datetime.strptime(datetime_str, "%Y/%m/%d %H:%M:%S")
    return datetime_obj.strftime("%m%d")

def process_excel(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
        cell = row[0]
        cell.value = convert_datetime_to_month_day(cell.value)

    output_path = file_path.replace(".xlsx", "_output.xlsx")
    wb.save(output_path)

def main():
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    file_path = filedialog.askopenfilename(title="Select Excel file", filetypes=[("Excel files", "*.xlsx")])

    if file_path:
        process_excel(file_path)
        print("Processing complete. Output saved as '_output.xlsx'.")

if __name__ == "__main__":
    main()
