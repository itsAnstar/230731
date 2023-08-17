import tkinter as tk
from tkinter import filedialog
import openpyxl
import pyautogui
import time
import pyperclip

# 全局变量
points = []
count = 0

# 选择文件
file_path = filedialog.askopenfilename() 

# 读取Excel
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

# 复制内容
text = sheet.cell(row=2, column=1).value
pyperclip.copy(text)

# 创建窗口
root = tk.Tk()

# 设置窗口大小
root.minsize(300, 200)

# 使用grid布局 
root.grid_rowconfigure(0, weight=1)
root.grid_columnconfigure(0, weight=1)

# 标签使用grid布局
label1 = tk.Label(root, bg='gray', font=('Arial', 12))
label1.grid(row=0, column=0, pady=5, padx=5)

# 按钮使用grid布局
btn_select = tk.Button(root, text='选择坐标', font=('Arial', 12), width=10)  
btn_select.grid(row=1, column=0, pady=5, padx=5)

btn_start = tk.Button(root, text='开始任务', font=('Arial', 12), width=10)
btn_start.grid(row=3, column=0, pady=5, padx=5)

# 点击函数
def on_click(event):

  global count
  if count < 5:

    point = pyautogui.position()
    points.append(point)

    if count == 0:
        label1.config(text=str(point), bg='pink') 

    elif count == 1:
        label2.config(text=str(point), bg='pink')

    count += 1

  if count == 5:
    btn_select['state'] = 'disabled'


# 绑定事件  
root.bind('<Button-1>', on_click)

# 创建按钮
btn_select = tk.Button(root, text='选择坐标') 
btn_start = tk.Button(root, text='开始任务')

# 开始任务
def start_task(btn_start):
  
  if len(points) < 5:
    print('选择坐标')
    return

  for row in range(2, sheet.max_row+1):

    # 点击坐标
    pyautogui.click(points[0])  

    # 检查键盘
    while True:
      if pyautogui.press('space'):
        break
          
  # 操作外部按钮实例     
  btn_start.config(text='完成')

# 调用函数  
start_task(btn_start)


root.mainloop()